from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.utils.text import slugify
from django.utils import timezone
from django.http import HttpResponse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import AdmissionApplication, ContactInquiry, Achievement, Event, Course, Blog, Testimonial, HomePageVideo, TeamMember, Brochure, Notification
from .serializers import AdmissionApplicationSerializer, ContactInquirySerializer

def index(request):
    testimonials = Testimonial.objects.filter(is_deleted=False).order_by('-created_at')
    video = HomePageVideo.objects.first()
    blogs = Blog.objects.filter(is_deleted=False).order_by('-created_at')[:3]
    notifications = Notification.objects.filter(is_active=True).order_by('-created_at')
    return render(request, 'admission/index.html', {
        'testimonials': testimonials, 
        'video': video, 
        'blogs': blogs,
        'notifications': notifications
    })

def about(request):
    team = TeamMember.objects.filter(is_deleted=False)
    return render(request, 'admission/about.html', {'team': team})

def courses(request):
    return render(request, 'admission/courses.html')

def contact(request):
    return render(request, 'admission/contact.html')

def register(request):
    return render(request, 'admission/register.html')

def blog(request):
    blogs_list = Blog.objects.filter(is_deleted=False).order_by('-created_at')
    return render(request, 'admission/blog.html', {'blogs': blogs_list})

def gallery(request):
    achievements = Achievement.objects.filter(is_deleted=False).order_by('-created_at')
    events = Event.objects.filter(is_deleted=False).order_by('-created_at')
    return render(request, 'admission/gallery.html', {
        'achievements': achievements,
        'events': events
    })

def course_detail_alm(request):
    return render(request, 'admission/alm.html')

def course_detail_bayan(request):
    return render(request, 'admission/bayan.html')

def course_detail_nuzul(request):
    return render(request, 'admission/nuzul.html')

def hifz(request):
    return render(request, 'admission/hifz.html')

def crash_course(request):
    return render(request, 'admission/crash-course.html')

# --- Admin Dashboard Views ---

@login_required
def admin_dashboard(request):
    from django.db.models import Count
    from django.db.models.functions import TruncDate
    from django.utils import timezone
    import json

    admissions = AdmissionApplication.objects.filter(is_deleted=False).order_by('-created_at')
    inquiries = ContactInquiry.objects.filter(is_deleted=False).order_by('-created_at')
    courses = Course.objects.filter(is_deleted=False)
    blogs = Blog.objects.filter(is_deleted=False)
    achievements = Achievement.objects.filter(is_deleted=False)
    events = Event.objects.filter(is_deleted=False)
    testimonials = Testimonial.objects.filter(is_deleted=False)
    upcoming_events = Event.objects.filter(is_deleted=False, date__gte=timezone.now().date()).order_by('date')[:5]

    # --- Chart: Admissions by course ---
    adm_by_course = (AdmissionApplication.objects
        .filter(is_deleted=False)
        .exclude(course__isnull=True).exclude(course='')
        .values('course').annotate(count=Count('id')).order_by('-count'))
    adm_course_labels = json.dumps([x['course'] for x in adm_by_course])
    adm_course_data   = json.dumps([x['count']  for x in adm_by_course])

    # --- Chart: Inquiries by course ---
    inq_by_course = (ContactInquiry.objects
        .filter(is_deleted=False)
        .exclude(course_interest__isnull=True).exclude(course_interest='')
        .values('course_interest').annotate(count=Count('id')).order_by('-count'))
    inq_course_labels = json.dumps([x['course_interest'] for x in inq_by_course])
    inq_course_data   = json.dumps([x['count']           for x in inq_by_course])

    # --- Chart: Daily counts last 30 days ---
    thirty_days_ago = timezone.now() - timezone.timedelta(days=30)

    adm_daily = (AdmissionApplication.objects
        .filter(is_deleted=False, created_at__gte=thirty_days_ago)
        .annotate(day=TruncDate('created_at'))
        .values('day').annotate(count=Count('id')).order_by('day'))

    inq_daily = (ContactInquiry.objects
        .filter(is_deleted=False, created_at__gte=thirty_days_ago)
        .annotate(day=TruncDate('created_at'))
        .values('day').annotate(count=Count('id')).order_by('day'))

    # Build a merged date set
    date_set = sorted(set(
        [x['day'].strftime('%b %d') for x in adm_daily] +
        [x['day'].strftime('%b %d') for x in inq_daily]
    ))
    adm_daily_map = {x['day'].strftime('%b %d'): x['count'] for x in adm_daily}
    inq_daily_map = {x['day'].strftime('%b %d'): x['count'] for x in inq_daily}

    daily_labels   = json.dumps(date_set)
    daily_adm_data = json.dumps([adm_daily_map.get(d, 0) for d in date_set])
    daily_inq_data = json.dumps([inq_daily_map.get(d, 0) for d in date_set])

    stats = {
        'total_admissions': admissions.count(),
        'total_inquiries': inquiries.count(),
        'total_courses': courses.count(),
        'total_blogs': blogs.count(),
        'total_achievements': achievements.count(),
        'total_events': events.count(),
        'total_testimonials': testimonials.count(),
        'recent_admissions': admissions[:5],
        'recent_inquiries': inquiries[:5],
    }

    return render(request, 'manager/dashboard.html', {
        'stats': stats,
        'upcoming_events': upcoming_events,
        'adm_course_labels': adm_course_labels,
        'adm_course_data': adm_course_data,
        'inq_course_labels': inq_course_labels,
        'inq_course_data': inq_course_data,
        'daily_labels': daily_labels,
        'daily_adm_data': daily_adm_data,
        'daily_inq_data': daily_inq_data,
    })


@login_required
def admin_admissions(request):
    selected_course = request.GET.get('course', '')
    selected_status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    admissions_list = AdmissionApplication.objects.filter(is_deleted=False).order_by('-created_at')
    
    if selected_course:
        admissions_list = admissions_list.filter(course__icontains=selected_course)
    if selected_status:
        admissions_list = admissions_list.filter(status=selected_status)
    if date_from:
        admissions_list = admissions_list.filter(created_at__date__gte=date_from)
    if date_to:
        admissions_list = admissions_list.filter(created_at__date__lte=date_to)
    
    paginator = Paginator(admissions_list, 10)
    page_number = request.GET.get('page')
    admissions = paginator.get_page(page_number)
        
    return render(request, 'manager/dashboard_admissions.html', {
        'admissions': admissions,
        'selected_course': selected_course,
        'selected_status': selected_status,
        'date_from': date_from,
        'date_to': date_to
    })

@login_required
def export_admissions_excel(request):
    selected_course = request.GET.get('course', '')
    selected_status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    admissions_list = AdmissionApplication.objects.filter(is_deleted=False).order_by('-created_at')
    
    if selected_course:
        admissions_list = admissions_list.filter(course__icontains=selected_course)
    if selected_status:
        admissions_list = admissions_list.filter(status=selected_status)
    if date_from:
        admissions_list = admissions_list.filter(created_at__date__gte=date_from)
    if date_to:
        admissions_list = admissions_list.filter(created_at__date__lte=date_to)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Admission Applications"
    
    # Headers
    headers = ['ID', 'First Name', 'Last Name', 'Gender', 'Age', 'Mobile', 'Email', 'Course', 'Status', 'Date']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        
    for app in admissions_list:
        ws.append([
            app.pk, app.first_name, app.last_name, app.gender, app.age, 
            app.mobile_number, app.email or '—', app.course or '—', app.status, 
            app.created_at.strftime("%Y-%m-%d %H:%M")
        ])
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Admissions_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    return response

@login_required
def mark_admission_viewed(request, pk):
    admission = get_object_or_404(AdmissionApplication, pk=pk)
    admission.status = 'viewed'
    admission.save()
    return redirect('admin_admissions')

@login_required
def unmark_admission_pending(request, pk):
    admission = get_object_or_404(AdmissionApplication, pk=pk)
    admission.status = 'pending'
    admission.save()
    return redirect('admin_admissions')

@login_required
def soft_delete_admission(request, pk):
    admission = get_object_or_404(AdmissionApplication, pk=pk)
    admission.is_deleted = True
    admission.save()
    return redirect('admin_admissions')

@login_required
def restore_admission(request, pk):
    admission = get_object_or_404(AdmissionApplication, pk=pk)
    admission.is_deleted = False
    admission.save()
    return redirect('recycle_bin_admissions')

@login_required
def recycle_bin_admissions(request):
    admissions_list = AdmissionApplication.objects.filter(is_deleted=True).order_by('-created_at')
    return render(request, 'manager/recycle_bin_admissions.html', {
        'admissions': admissions_list
    })

@login_required
def delete_admission_permanent(request, pk):
    admission = get_object_or_404(AdmissionApplication, pk=pk)
    admission.delete()
    return redirect('recycle_bin_admissions')

@login_required
def admin_inquiries(request):
    selected_course = request.GET.get('course', '')
    selected_status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    inquiries_list = ContactInquiry.objects.filter(is_deleted=False).order_by('-created_at')
    
    if selected_course:
        inquiries_list = inquiries_list.filter(course_interest__icontains=selected_course)
    if selected_status:
        inquiries_list = inquiries_list.filter(status=selected_status)
    if date_from:
        inquiries_list = inquiries_list.filter(created_at__date__gte=date_from)
    if date_to:
        inquiries_list = inquiries_list.filter(created_at__date__lte=date_to)
    
    paginator = Paginator(inquiries_list, 10)
    page_number = request.GET.get('page')
    inquiries = paginator.get_page(page_number)
        
    return render(request, 'manager/dashboard_inquiries.html', {
        'inquiries': inquiries,
        'selected_course': selected_course,
        'selected_status': selected_status,
        'date_from': date_from,
        'date_to': date_to
    })

@login_required
def export_inquiries_excel(request):
    selected_course = request.GET.get('course', '')
    selected_status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    inquiries_list = ContactInquiry.objects.filter(is_deleted=False).order_by('-created_at')
    
    if selected_course:
        inquiries_list = inquiries_list.filter(course_interest__icontains=selected_course)
    if selected_status:
        inquiries_list = inquiries_list.filter(status=selected_status)
    if date_from:
        inquiries_list = inquiries_list.filter(created_at__date__gte=date_from)
    if date_to:
        inquiries_list = inquiries_list.filter(created_at__date__lte=date_to)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contact Inquiries"
    
    # Headers
    headers = ['ID', 'Full Name', 'Email', 'Phone', 'Course Interest', 'Message', 'Status', 'Date']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        
    for inq in inquiries_list:
        ws.append([
            inq.pk, inq.full_name, inq.email or '—', inq.phone_number, 
            inq.course_interest or 'General', inq.message, inq.status, 
            inq.created_at.strftime("%Y-%m-%d %H:%M")
        ])
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Inquiries_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    return response

@login_required
def mark_inquiry_viewed(request, pk):
    inquiry = get_object_or_404(ContactInquiry, pk=pk)
    inquiry.status = 'viewed'
    inquiry.save()
    return redirect('admin_inquiries')

@login_required
def unmark_inquiry_pending(request, pk):
    inquiry = get_object_or_404(ContactInquiry, pk=pk)
    inquiry.status = 'pending'
    inquiry.save()
    return redirect('admin_inquiries')

@login_required
def soft_delete_inquiry(request, pk):
    inquiry = get_object_or_404(ContactInquiry, pk=pk)
    inquiry.is_deleted = True
    inquiry.save()
    return redirect('admin_inquiries')

@login_required
def restore_inquiry(request, pk):
    inquiry = get_object_or_404(ContactInquiry, pk=pk)
    inquiry.is_deleted = False
    inquiry.save()
    return redirect('recycle_bin_inquiries')

@login_required
def recycle_bin_inquiries(request):
    inquiries_list = ContactInquiry.objects.filter(is_deleted=True).order_by('-created_at')
    return render(request, 'manager/recycle_bin_inquiries.html', {
        'inquiries': inquiries_list
    })

@login_required
def delete_inquiry_permanent(request, pk):
    inquiry = get_object_or_404(ContactInquiry, pk=pk)
    inquiry.delete()
    return redirect('recycle_bin_inquiries')

@login_required
def admin_achievements(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        image = request.FILES.get('image')
        if title and image:
            Achievement.objects.create(title=title, image=image)
            return redirect('admin_achievements')
            
    achievements = Achievement.objects.filter(is_deleted=False).order_by('-created_at')
    return render(request, 'manager/dashboard_achievements.html', {
        'achievements': achievements
    })

@login_required
def soft_delete_achievement(request, pk):
    achievement = get_object_or_404(Achievement, pk=pk)
    achievement.is_deleted = True
    achievement.save()
    return redirect('admin_achievements')

@login_required
def restore_achievement(request, pk):
    achievement = get_object_or_404(Achievement, pk=pk)
    achievement.is_deleted = False
    achievement.save()
    return redirect('recycle_bin_achievements')

@login_required
def recycle_bin_achievements(request):
    achievements_list = Achievement.objects.filter(is_deleted=True).order_by('-created_at')
    return render(request, 'manager/recycle_bin_achievements.html', {
        'achievements': achievements_list
    })

@login_required
def delete_achievement_permanent(request, pk):
    achievement = get_object_or_404(Achievement, pk=pk)
    achievement.delete()
    return redirect('recycle_bin_achievements')

@login_required
def edit_achievement(request, pk):
    achievement = get_object_or_404(Achievement, pk=pk)
    if request.method == 'POST':
        achievement.title = request.POST.get('title')
        if request.FILES.get('image'):
            achievement.image = request.FILES.get('image')
        achievement.save()
        return redirect('admin_achievements')
    return render(request, 'manager/edit_item.html', {
        'item': achievement,
        'type': 'Achievement',
        'title_label': 'Achievement Title',
        'back_url': 'admin_achievements'
    })

@login_required
def admin_events(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        image = request.FILES.get('image')
        date = request.POST.get('date')
        if title and image:
            Event.objects.create(title=title, image=image, date=date or None)
            return redirect('admin_events')
            
    events = Event.objects.filter(is_deleted=False).order_by('-created_at')
    return render(request, 'manager/dashboard_events.html', {
        'events': events
    })

@login_required
def soft_delete_event(request, pk):
    event = get_object_or_404(Event, pk=pk)
    event.is_deleted = True
    event.save()
    return redirect('admin_events')

@login_required
def restore_event(request, pk):
    event = get_object_or_404(Event, pk=pk)
    event.is_deleted = False
    event.save()
    return redirect('recycle_bin_events')

@login_required
def recycle_bin_events(request):
    events_list = Event.objects.filter(is_deleted=True).order_by('-created_at')
    return render(request, 'manager/recycle_bin_events.html', {
        'events': events_list
    })

@login_required
def delete_event_permanent(request, pk):
    event = get_object_or_404(Event, pk=pk)
    event.delete()
    return redirect('recycle_bin_events')

@login_required
def edit_event(request, pk):
    event = get_object_or_404(Event, pk=pk)
    if request.method == 'POST':
        event.title = request.POST.get('title')
        event.date = request.POST.get('date') or None
        if request.FILES.get('image'):
            event.image = request.FILES.get('image')
        event.save()
        return redirect('admin_events')
    return render(request, 'manager/edit_item.html', {
        'item': event,
        'type': 'Event',
        'title_label': 'Event Title',
        'has_date': True,
        'back_url': 'admin_events'
    })

@login_required
def admin_courses(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        image = request.FILES.get('image')
        if name and description:
            course_obj = Course(name=name, description=description, slug=slugify(name))
            if image:
                course_obj.image = image
            course_obj.save()
            return redirect('admin_courses')
            
    courses_list = Course.objects.filter(is_deleted=False).order_by('-created_at')
    return render(request, 'manager/dashboard_courses.html', {
        'courses': courses_list
    })

@login_required
def soft_delete_course(request, pk):
    course = get_object_or_404(Course, pk=pk)
    course.is_deleted = True
    course.save()
    return redirect('admin_courses')

@login_required
def restore_course(request, pk):
    course = get_object_or_404(Course, pk=pk)
    course.is_deleted = False
    course.save()
    return redirect('recycle_bin_courses')

@login_required
def recycle_bin_courses(request):
    courses_list = Course.objects.filter(is_deleted=True).order_by('-created_at')
    return render(request, 'manager/recycle_bin_courses.html', {
        'courses': courses_list
    })

@login_required
def delete_course_permanent(request, pk):
    course = get_object_or_404(Course, pk=pk)
    course.delete()
    return redirect('recycle_bin_courses')

@login_required
def edit_course(request, pk):
    course = get_object_or_404(Course, pk=pk)
    if request.method == 'POST':
        course.name = request.POST.get('name')
        course.description = request.POST.get('description')
        if request.FILES.get('image'):
            course.image = request.FILES.get('image')
        course.slug = slugify(course.name)
        course.save()
        return redirect('admin_courses')
    return render(request, 'manager/edit_item.html', {
        'item': course,
        'type': 'Course',
        'title_label': 'Course Name',
        'has_description': True,
        'back_url': 'admin_courses'
    })

@login_required
def admin_blogs(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        content_p1 = request.POST.get('content_p1')
        quote = request.POST.get('quote')
        content_p2 = request.POST.get('content_p2')
        thumbnail = request.FILES.get('image')
        image = request.FILES.get('image')
        
        if title and content_p1:
            Blog.objects.create(
                title=title,
                content_p1=content_p1,
                quote=quote,
                content_p2=content_p2,
                thumbnail=thumbnail,
                image=image,
                slug=slugify(title)
            )
            return redirect('admin_blogs')
            
    blogs_list = Blog.objects.filter(is_deleted=False).order_by('-created_at')
    return render(request, 'manager/dashboard_blogs.html', {
        'blogs': blogs_list
    })

@login_required
def soft_delete_blog(request, pk):
    blog = get_object_or_404(Blog, pk=pk)
    blog.is_deleted = True
    blog.save()
    return redirect('admin_blogs')

@login_required
def restore_blog(request, pk):
    blog = get_object_or_404(Blog, pk=pk)
    blog.is_deleted = False
    blog.save()
    return redirect('recycle_bin_blogs')

@login_required
def recycle_bin_blogs(request):
    blogs_list = Blog.objects.filter(is_deleted=True).order_by('-created_at')
    return render(request, 'manager/recycle_bin_blogs.html', {
        'blogs': blogs_list
    })

@login_required
def delete_blog_permanent(request, pk):
    blog = get_object_or_404(Blog, pk=pk)
    blog.delete()
    return redirect('recycle_bin_blogs')

@login_required
def edit_blog(request, pk):
    blog = get_object_or_404(Blog, pk=pk)
    if request.method == 'POST':
        blog.title = request.POST.get('title')
        blog.content_p1 = request.POST.get('content_p1')
        blog.quote = request.POST.get('quote')
        blog.content_p2 = request.POST.get('content_p2')
        
        if request.FILES.get('image'):
            img = request.FILES.get('image')
            blog.thumbnail = img
            blog.image = img
            
        blog.slug = slugify(blog.title)
        blog.save()
        return redirect('admin_blogs')
        
    return render(request, 'manager/edit_blog.html', {
        'blog': blog,
        'back_url': 'admin_blogs'
    })

def blog_detail(request, slug):
    blog = get_object_or_404(Blog, slug=slug, is_deleted=False)
    recent_blogs = Blog.objects.filter(is_deleted=False).exclude(id=blog.id).order_by('-created_at')[:3]
    return render(request, 'admission/blog_detail.html', {
        'blog': blog,
        'recent_blogs': recent_blogs
    })

# --- API View Handlers ---

@api_view(['POST'])
def submit_admission(request):
    serializer = AdmissionApplicationSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response({'message': 'Application submitted!'}, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
def submit_contact(request):
    serializer = ContactInquirySerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response({'message': 'Message sent successfully!'}, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@login_required
def admin_testimonials(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        role = request.POST.get('role')
        content = request.POST.get('content')
        image = request.FILES.get('image')
        if name and content and image:
            Testimonial.objects.create(name=name, role=role, content=content, image=image)
            return redirect('admin_testimonials')
            
    testimonials = Testimonial.objects.filter(is_deleted=False).order_by('-created_at')
    return render(request, 'manager/dashboard_testimonials.html', {
        'testimonials': testimonials
    })

@login_required
def soft_delete_testimonial(request, pk):
    testimonial = get_object_or_404(Testimonial, pk=pk)
    testimonial.is_deleted = True
    testimonial.save()
    return redirect('admin_testimonials')

@login_required
def restore_testimonial(request, pk):
    testimonial = get_object_or_404(Testimonial, pk=pk)
    testimonial.is_deleted = False
    testimonial.save()
    return redirect('recycle_bin_testimonials')

@login_required
def recycle_bin_testimonials(request):
    testimonials_list = Testimonial.objects.filter(is_deleted=True).order_by('-created_at')
    return render(request, 'manager/recycle_bin_testimonials.html', {
        'testimonials': testimonials_list
    })

@login_required
def delete_testimonial_permanent(request, pk):
    testimonial = get_object_or_404(Testimonial, pk=pk)
    testimonial.delete()
    return redirect('recycle_bin_testimonials')

@login_required
def edit_testimonial(request, pk):
    testimonial = get_object_or_404(Testimonial, pk=pk)
    if request.method == 'POST':
        testimonial.name = request.POST.get('name')
        testimonial.role = request.POST.get('role')
        testimonial.content = request.POST.get('content')
        if request.FILES.get('image'):
            testimonial.image = request.FILES.get('image')
        testimonial.save()
        return redirect('admin_testimonials')
    return render(request, 'manager/edit_item.html', {
        'item': testimonial,
        'type': 'Testimonial',
        'title_label': 'Student Name',
        'has_content': True,
        'has_role': True,
        'back_url': 'admin_testimonials'
    })

@login_required
def admin_video(request):
    video = HomePageVideo.objects.first()
    if request.method == 'POST':
        url = request.POST.get('youtube_url')
        if video:
            video.youtube_url = url
            video.save()
        else:
            video = HomePageVideo.objects.create(youtube_url=url)
        return redirect('admin_video')
    
    return render(request, 'manager/dashboard_video.html', {'video': video})

@login_required
def admin_team(request):
    team_list = TeamMember.objects.filter(is_deleted=False).order_by('-created_at')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        role = request.POST.get('role')
        image = request.FILES.get('image')
        
        TeamMember.objects.create(name=name, role=role, image=image)
        return redirect('admin_team')
        
    return render(request, 'manager/dashboard_team.html', {
        'team_list': team_list
    })

@login_required
def soft_delete_team(request, pk):
    member = get_object_or_404(TeamMember, pk=pk)
    member.is_deleted = True
    member.save()
    return redirect('admin_team')

@login_required
def edit_team(request, pk):
    member = get_object_or_404(TeamMember, pk=pk)
    if request.method == 'POST':
        member.name = request.POST.get('name')
        member.role = request.POST.get('role')
        if request.FILES.get('image'):
            member.image = request.FILES.get('image')
        member.save()
        return redirect('admin_team')
    return render(request, 'manager/edit_item.html', {
        'item': member,
        'type': 'Team',
        'title_label': 'Member Name',
        'has_content': False,
        'has_role': True,
        'back_url': 'admin_team'
    })

@login_required
def recycle_bin_team(request):
    team_list = TeamMember.objects.filter(is_deleted=True).order_by('-created_at')
    return render(request, 'manager/recycle_bin_team.html', {
        'team_list': team_list
    })

@login_required
def restore_team(request, pk):
    member = get_object_or_404(TeamMember, pk=pk)
    member.is_deleted = False
    member.save()
    return redirect('recycle_bin_team')

@login_required
def delete_team_permanent(request, pk):
    member = get_object_or_404(TeamMember, pk=pk)
    member.delete()
    return redirect('recycle_bin_team')

@login_required
def admin_brochure(request):
    brochure = Brochure.objects.last()
    if request.method == 'POST':
        file = request.FILES.get('brochure_file')
        if file:
            if brochure:
                brochure.file = file
                brochure.save()
            else:
                brochure = Brochure.objects.create(file=file)
        return redirect('admin_brochure')
    
    return render(request, 'manager/dashboard_brochure.html', {'brochure': brochure})
@login_required
def admin_notifications(request):
    if request.method == 'POST':
        text = request.POST.get('text')
        link = request.POST.get('link')
        if text:
            Notification.objects.create(text=text, link=link)
            return redirect('admin_notifications')
            
    notifications = Notification.objects.all().order_by('-created_at')
    return render(request, 'manager/dashboard_notifications.html', {
        'notifications': notifications
    })

@login_required
def toggle_notification(request, pk):
    notif = get_object_or_404(Notification, pk=pk)
    notif.is_active = not notif.is_active
    notif.save()
    return redirect('admin_notifications')

@login_required
def delete_notification(request, pk):
    notif = get_object_or_404(Notification, pk=pk)
    notif.delete()
    return redirect('admin_notifications')

@login_required
def edit_notification(request, pk):
    notif = get_object_or_404(Notification, pk=pk)
    if request.method == 'POST':
        notif.text = request.POST.get('text')
        notif.link = request.POST.get('link')
        notif.save()
        return redirect('admin_notifications')
    return render(request, 'manager/edit_item.html', {
        'item': notif,
        'type': 'Notification',
        'title_label': 'Notification Text',
        'has_link': True,
        'back_url': 'admin_notifications'
    })


# Error Views

def custom_400(request, exception):
    context = {
        'error_code': 400,
        'error_description': 'This request could not be processed.',
    }
    return render(request, 'errors/error_page.html', context=context, status=400)


def custom_403(request, exception):
    context = {
        'error_code': 403,
        'error_description': 'You are not allowed to view this page.',
    }
    return render(request, 'errors/error_page.html', context=context, status=403)


def custom_404(request, exception):
    context = {
        'error_code': 404,
        'error_description': "This page isn't here",
    }
    return render(request, 'errors/error_page.html', context=context, status=404)


def custom_500(request):
    context = {
        'error_code': 500,
        'error_description': 'Something went wrong. Please try again later.',
    }
    return render(request, 'errors/error_page.html', context=context, status=500)



    # Arabic Site Template Views
def arabic_learn_for_kids_index(request):
    return render(request, 'arabic/arabic_learn_for_kids.html')

def arabic_learn_ustad_index(request):
    return render(request, 'arabic/arabic_learn_ustad.html')